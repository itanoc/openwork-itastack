#!/usr/bin/env python3
"""write_scope.py — Write task plan data to a copy of the Sales Scope Template.

Reads a JSON task plan from stdin or a file and populates the Tasklist tab:
  - Column A: Task descriptions (organized by day)
  - Column C: Time estimates (minimum hours)
  - Column G: Downtime? (Yes/No)
  - Column H: Afterhours? (Yes/No)
  - Row 2 metadata: Project name, ticket number, prepared by, date
  - Downtime explanations, parts, client/vendor dependencies, comments
  - Color coding per task type (remote=blue, onsite=purple, vendor=red,
    client=yellow, procurement=green)

Preserves all existing formulas, formatting, and other tabs.
Handles dynamic day counts by inserting rows when tasks exceed the
template's default 4-day/25-row capacity.
Creates automatic backup before writing.

Usage:
  python write_scope.py --input plan.json --file /path/to/scope-copy.xlsx
  cat plan.json | python write_scope.py --file /path/to/scope-copy.xlsx
  python write_scope.py --input plan.json --file scope.xlsx --no-backup

JSON format:
{
  "metadata": {
    "project_name": "RMCN - New Server",
    "ticket_number": "59542",
    "prepared_by": "Riely Borek",
    "date": "2026-03-18"
  },
  "days": [
    {
      "label": "Day 1 - 8hrs max",
      "tasks": [
        {
          "description": "Rack and cable new server hardware",
          "time_min": 1.5,
          "downtime": "No",
          "afterhours": "No",
          "location": "onsite"
        }
      ]
    }
  ],
  "downtime_explanation": "VM migration requires shutting down VMs...",
  "parts": [
    {
      "description": "Dell PowerEdge R760xs",
      "quantity": 1,
      "part_number": "R760XS-001",
      "url": "https://dell.com/...",
      "price": 5500.00,
      "alternative": "HPE ProLiant DL360"
    }
  ],
  "client_dependencies": "Client must provide downtime approval 48hrs in advance.",
  "cat_herding": "No",
  "vendor_dependencies": {
    "vendors": [
      {"name": "Dell ProSupport"},
      {"name": "ISP - Comcast"}
    ],
    "support_current": "Yes",
    "vendor_charges": "No",
    "existing_ticket": "",
    "contact": "1-800-456-3355",
    "hours": "24/7"
  },
  "comments": "Additional notes..."
}
"""

import argparse
import json
import shutil
import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# --- Template fixed row positions (before any row insertion) ---
TASK_START_ROW = 5
DEFAULT_TASK_END_ROW = 29  # Template provides rows 5-29 for tasks

# --- Color coding fills ---
# Template convention:
#   Blue = ITA_office/Remote
#   Purple = ITA_client onsite
#   Red = Vendor Responsibility
#   Yellow = Client Responsibility
#   Green = Procurement
COLOR_MAP = {
    "remote":      PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),  # Light blue
    "onsite":      PatternFill(start_color="D5A6E6", end_color="D5A6E6", fill_type="solid"),  # Light purple
    "vendor":      PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),  # Light red
    "client":      PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Light yellow
    "procurement": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),  # Light green
}

# Day header style
DAY_HEADER_FONT = Font(bold=True, size=11)
DAY_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray


def load_plan(input_path=None):
    """Load task plan from file or stdin."""
    if input_path:
        with open(input_path, 'r') as f:
            return json.load(f)
    else:
        return json.load(sys.stdin)


def count_task_rows(plan):
    """Count total rows needed for all days (day headers + tasks)."""
    total = 0
    for day in plan.get('days', []):
        total += 1  # Day header row
        total += len(day.get('tasks', []))
    return total


def copy_row_style(ws, source_row, target_row, max_col=9):
    """Copy cell formatting from source_row to target_row."""
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=source_row, column=col)
        tgt_cell = ws.cell(row=target_row, column=col)
        if src_cell.has_style:
            tgt_cell.font = src_cell.font.copy()
            tgt_cell.border = src_cell.border.copy()
            tgt_cell.fill = src_cell.fill.copy()
            tgt_cell.number_format = src_cell.number_format
            tgt_cell.alignment = src_cell.alignment.copy()
            tgt_cell.protection = src_cell.protection.copy()


def apply_task_color(ws, row, location, max_col=9):
    """Apply color fill to Column A only for a task row based on location type."""
    if not location:
        return
    fill = COLOR_MAP.get(location.lower())
    if fill:
        ws.cell(row=row, column=1).fill = fill  # Column A only


def apply_day_header_style(ws, row, max_col=9):
    """Apply bold gray style to day header rows."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DAY_HEADER_FONT
        cell.fill = DAY_HEADER_FILL


def create_backup(xlsx_path):
    """Create a timestamped backup of the file before modifying."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(xlsx_path)
    backup_path = f"{base}_backup_{timestamp}{ext}"
    shutil.copy2(xlsx_path, backup_path)
    print(f"  Backup created: {backup_path}")
    return backup_path


def write_scope(plan, xlsx_path, no_backup=False):
    """Write the task plan into the Tasklist tab of the scope template."""
    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    # --- Create backup ---
    if not no_backup:
        create_backup(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)

    if 'Tasklist' not in wb.sheetnames:
        print("ERROR: 'Tasklist' sheet not found in workbook")
        sys.exit(1)

    ws = wb['Tasklist']

    # --- Calculate row needs ---
    rows_needed = count_task_rows(plan)
    rows_available = DEFAULT_TASK_END_ROW - TASK_START_ROW + 1  # 25 rows
    extra_rows = max(0, rows_needed - rows_available)

    # Capture style from a representative task row before insertion
    style_source_row = TASK_START_ROW + 1  # Row 6 is a typical task row

    # Insert extra rows if needed
    if extra_rows > 0:
        ws.insert_rows(DEFAULT_TASK_END_ROW + 1, amount=extra_rows)
        # Apply formatting to inserted rows
        for i in range(extra_rows):
            new_row = DEFAULT_TASK_END_ROW + 1 + i
            copy_row_style(ws, style_source_row, new_row)
        print(f"  Inserted {extra_rows} extra rows to accommodate {rows_needed} task rows")

    actual_task_end_row = DEFAULT_TASK_END_ROW + extra_rows
    offset = extra_rows

    # --- Write metadata (Row 2) ---
    meta = plan.get('metadata', {})
    if meta.get('project_name'):
        ws['A2'] = f"Project Name: {meta['project_name']}"
    if meta.get('ticket_number'):
        ws['B2'] = f"Project Ticket Number: {meta['ticket_number']}"
    if meta.get('prepared_by'):
        ws['C2'] = f"Sales Scope Prepared by: {meta['prepared_by']}"
    if meta.get('date'):
        ws['G2'] = meta['date']

    # --- Clear existing task content ---
    for row in range(TASK_START_ROW, actual_task_end_row + 1):
        for col in ['A', 'B', 'C', 'G', 'H']:
            ws[f'{col}{row}'] = None
        # Reset fills on task rows
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

    # --- Write task rows ---
    days = plan.get('days', [])
    current_row = TASK_START_ROW
    day_hours = {}

    for day_idx, day in enumerate(days):
        # Write day header
        day_label = day.get('label', f'Day {day_idx + 1} - 8hrs max')
        ws[f'A{current_row}'] = day_label
        ws[f'D{current_row}'] = f'=(C{current_row}+E{current_row})/2'
        ws[f'E{current_row}'] = f'=C{current_row}*2'
        apply_day_header_style(ws, current_row)
        current_row += 1

        # Write tasks for this day
        tasks = day.get('tasks', [])
        total_hours = 0
        for task in tasks:
            ws[f'A{current_row}'] = task.get('description', '')
            time_min = task.get('time_min')
            if time_min is not None:
                ws[f'C{current_row}'] = time_min
                total_hours += time_min
            if task.get('downtime'):
                ws[f'G{current_row}'] = task['downtime']
            if task.get('afterhours'):
                ws[f'H{current_row}'] = task['afterhours']
            ws[f'D{current_row}'] = f'=(C{current_row}+E{current_row})/2'
            ws[f'E{current_row}'] = f'=C{current_row}*2'

            # Apply color coding
            location = task.get('location', '')
            if location:
                apply_task_color(ws, current_row, location)

            current_row += 1

        day_hours[day_label] = total_hours
        if total_hours > 8:
            print(f"  \u26a0\ufe0f  WARNING: {day_label} totals {total_hours}hrs (exceeds 8hr max)")

    # --- Update fixed sections below tasks ---
    at_completion_row = 30 + offset
    ws[f'A{at_completion_row}'] = "At completion of project:"

    ws[f'A{at_completion_row + 1}'] = "Update ITGlue"
    ws[f'D{at_completion_row + 1}'] = f'=(C{at_completion_row + 1}+E{at_completion_row + 1})/2'
    ws[f'E{at_completion_row + 1}'] = f'=C{at_completion_row + 1}*2'

    ws[f'A{at_completion_row + 2}'] = "Deprecate old ITGlue entries"
    ws[f'D{at_completion_row + 2}'] = f'=(C{at_completion_row + 2}+E{at_completion_row + 2})/2'
    ws[f'E{at_completion_row + 2}'] = f'=C{at_completion_row + 2}*2'

    ws[f'A{at_completion_row + 3}'] = (
        "Notify Project Coordinator if any task was not performed "
        "or if the project did not achieve its stated goal"
    )

    # --- Update formulas ---
    last_task_row = actual_task_end_row
    completion_last_row = at_completion_row + 3

    dt_est_row = 35 + offset
    ws[f'A{dt_est_row}'] = "Downtime Estimates"

    dt_total_row = 36 + offset
    ws[f'A{dt_total_row}'] = "Estimated Downtime Total:"
    ws[f'C{dt_total_row}'] = f'=SUMIF(G{TASK_START_ROW}:G{last_task_row},"Yes",C{TASK_START_ROW}:C{last_task_row})'
    ws[f'D{dt_total_row}'] = f'=(C{dt_total_row}+E{dt_total_row})/2'
    ws[f'E{dt_total_row}'] = f'=C{dt_total_row}*2'

    ah_total_row = 37 + offset
    ws[f'A{ah_total_row}'] = "Recommended After Hours Labor"
    ws[f'C{ah_total_row}'] = f'=SUMIF(H{TASK_START_ROW}:H{last_task_row},"Yes",C{TASK_START_ROW}:C{last_task_row})'
    ws[f'D{ah_total_row}'] = f'=(C{ah_total_row}+E{ah_total_row})/2'
    ws[f'E{ah_total_row}'] = f'=C{ah_total_row}*2'

    labor_row = 38 + offset
    ws[f'B{labor_row}'] = "Labor Total"
    ws[f'C{labor_row}'] = f'=SUM(C{TASK_START_ROW}:C{completion_last_row})'
    ws[f'D{labor_row}'] = f'=SUM(D{TASK_START_ROW}:D{completion_last_row})'
    ws[f'E{labor_row}'] = f'=SUM(E{TASK_START_ROW}:E{completion_last_row})'

    trouble_row = 39 + offset
    ws[f'B{trouble_row}'] = "Troubleshooting (25%)"
    ws[f'C{trouble_row}'] = f'=SUM(C{labor_row}*0.25)'
    ws[f'D{trouble_row}'] = f'=SUM(D{labor_row}*0.25)'
    ws[f'E{trouble_row}'] = f'=SUM(E{labor_row}*0.25)'

    total_row = 40 + offset
    ws[f'B{total_row}'] = "Total (Hours)"
    ws[f'C{total_row}'] = f'=SUM(C{labor_row}+C{trouble_row})'
    ws[f'D{total_row}'] = f'=SUM(D{labor_row}+D{trouble_row})'
    ws[f'E{total_row}'] = f'=SUM(E{labor_row}+E{trouble_row})'

    pm_row = 41 + offset
    ws[f'B{pm_row}'] = 'Proj Mgmt (20% of labor before "At completion of Project"'
    ws[f'C{pm_row}'] = f'=SUM(C{TASK_START_ROW}:C{last_task_row})*0.2'
    ws[f'D{pm_row}'] = f'=SUM(D{TASK_START_ROW}:D{last_task_row})*0.2'
    ws[f'E{pm_row}'] = f'=SUM(E{TASK_START_ROW}:E{last_task_row})*0.2'

    ws[f'B{42 + offset}'] = "Labor Quoted:"
    ws[f'B{43 + offset}'] = "Sales Notes:"

    # --- Write Downtime Explanation (rows 44-46+offset) ---
    dt_explain_row = 44 + offset
    ws[f'A{dt_explain_row}'] = "Downtime Explanation(s)"
    ws[f'A{dt_explain_row + 1}'] = (
        "If there will be downtime - what of the clients workflow will be affected?"
    )
    dt_explanation = plan.get('downtime_explanation', '')
    if dt_explanation:
        ws[f'A{dt_explain_row + 2}'] = dt_explanation
    else:
        ws[f'A{dt_explain_row + 2}'] = "Please be specific"

    # --- Write Parts Needed (rows 52+offset onward) ---
    parts_row = 52 + offset
    ws[f'A{parts_row}'] = (
        "Parts Needed: *Please add alternative options for all major hardware\n"
        "Please include any: Ethernet cables, velcro/zip ties, or other mounting hardware needed?"
    )
    ws[f'B{parts_row}'] = "Quantity"
    ws[f'C{parts_row}'] = "MFG part number:"
    ws[f'D{parts_row}'] = "URL:"
    ws[f'E{parts_row}'] = "Vendor Total QTY Price:"
    ws[f'F{parts_row}'] = "Hardware Alternative:"

    parts = plan.get('parts', [])
    for i, part in enumerate(parts):
        r = parts_row + 1 + i
        ws[f'A{r}'] = part.get('description', '')
        if part.get('quantity') is not None:
            ws[f'B{r}'] = part['quantity']
        if part.get('part_number'):
            ws[f'C{r}'] = part['part_number']
        if part.get('url'):
            ws[f'D{r}'] = part['url']
        if part.get('price') is not None:
            ws[f'E{r}'] = part['price']
        if part.get('alternative'):
            ws[f'F{r}'] = part['alternative']

    # Parts total row
    if parts:
        total_parts_row = parts_row + 1 + len(parts)
        ws[f'D{total_parts_row}'] = "Total:"
        ws[f'E{total_parts_row}'] = (
            f'=SUM(E{parts_row + 1}:E{parts_row + len(parts)})'
        )
    else:
        # Ensure empty price cells have 0 for existing template formula
        for i in range(1, 9):
            r = parts_row + i
            if ws[f'E{r}'].value is None:
                ws[f'E{r}'] = 0

    # --- Write Client Dependencies (rows 62+offset onward) ---
    client_dep_row = 62 + offset
    ws[f'A{client_dep_row}'] = "Client Dependencies:"
    client_deps = plan.get('client_dependencies', '')
    if client_deps:
        ws[f'A{client_dep_row + 1}'] = client_deps

    cat_herding_row = 63 + offset
    ws[f'A{cat_herding_row}'] = "Would this project involve cat-herding ? (Y/N):"
    cat_herding = plan.get('cat_herding', '')
    if cat_herding:
        ws[f'B{cat_herding_row}'] = cat_herding

    # --- Write Vendor Dependencies (rows 66+offset onward) ---
    vendor_dep_row = 66 + offset
    vendor_deps = plan.get('vendor_dependencies', {})
    vendors = vendor_deps.get('vendors', [])

    # Vendor names row
    vendor_name_row = vendor_dep_row + 1
    for i, v in enumerate(vendors[:3]):
        col = ['B', 'C', 'D'][i]
        ws[f'{col}{vendor_name_row}'] = f"Vendor {i+1}: {v.get('name', '')}"

    # Vendor detail rows
    ws[f'A{vendor_dep_row + 2}'] = "Is the clients support contract current:"
    if vendor_deps.get('support_current'):
        ws[f'B{vendor_dep_row + 2}'] = vendor_deps['support_current']

    ws[f'A{vendor_dep_row + 3}'] = "Does the vendor charge for the action you need help with:"
    if vendor_deps.get('vendor_charges'):
        ws[f'B{vendor_dep_row + 3}'] = vendor_deps['vendor_charges']

    ws[f'A{vendor_dep_row + 4}'] = "Is there already a ticket open with them? (Add #)"
    if vendor_deps.get('existing_ticket'):
        ws[f'B{vendor_dep_row + 4}'] = vendor_deps['existing_ticket']

    ws[f'A{vendor_dep_row + 5}'] = "Vendor contact:"
    if vendor_deps.get('contact'):
        ws[f'B{vendor_dep_row + 5}'] = vendor_deps['contact']

    ws[f'A{vendor_dep_row + 6}'] = "Business hours:"
    if vendor_deps.get('hours'):
        ws[f'B{vendor_dep_row + 6}'] = vendor_deps['hours']

    # --- Write Comments ---
    comments_row = 74 + offset
    ws[f'A{comments_row}'] = "Comments:"
    comments = plan.get('comments', '')
    if comments:
        ws[f'A{comments_row + 1}'] = comments

    # --- Save ---
    wb.save(xlsx_path)
    print(f"SUCCESS: Task plan written to {xlsx_path}")
    print(f"  Days: {len(days)}")
    total_tasks = sum(len(d.get('tasks', [])) for d in days)
    print(f"  Tasks: {total_tasks}")
    print(f"  Task rows: {TASK_START_ROW}-{current_row - 1}")
    if extra_rows > 0:
        print(f"  Extra rows inserted: {extra_rows}")
    if parts:
        print(f"  Parts: {len(parts)}")
    for label, hrs in day_hours.items():
        status = "\u2705" if hrs <= 8 else "\u26a0\ufe0f  OVER"
        print(f"  {label}: {hrs}hrs {status}")


def main():
    parser = argparse.ArgumentParser(
        description="Write scope task plan to Excel template"
    )
    parser.add_argument(
        '--input', '-i',
        help="Path to JSON plan file (default: stdin)"
    )
    parser.add_argument(
        '--file', '-f', required=True,
        help="Path to the Excel scope template copy"
    )
    parser.add_argument(
        '--no-backup', action='store_true',
        help="Skip creating a backup before writing"
    )
    args = parser.parse_args()

    plan = load_plan(args.input)
    write_scope(plan, args.file, no_backup=args.no_backup)


if __name__ == '__main__':
    main()
